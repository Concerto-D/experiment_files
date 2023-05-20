import yaml
from openpyxl import Workbook


def generate_sheet_expe_1(wb: Workbook, expe_name, data_path):
    col_mapping = {
        "2-2": {"synchronous": "C", "asynchronous": "D", "mjuz-2-comps": "B"},
        "25-35": {"synchronous": "F", "asynchronous": "G", "mjuz-2-comps": "E"},
        "50-50": {"synchronous": "I", "asynchronous": "J", "mjuz-2-comps": "H"},
        "1-1": {"synchronous": "L", "asynchronous": "M", "mjuz-2-comps": "K"},
    }

    def case_mapping(perc, categ):
        col = col_mapping[perc][categ]
        return {
            (perc, categ, "T0", "max_deploy_time"): f"{col}5",
            (perc, categ, "T0", "max_update_time"): f"{col}8",
            (perc, categ, "T1", "max_deploy_time"): f"{col}6",
            (perc, categ, "T1", "max_update_time"): f"{col}9",
            (perc, categ, "T0", "global_finished_reconf"): f"{col}20",
            (perc, categ, "T0", "files"): f"{col}21",
            (perc, categ, "T1", "global_finished_reconf"): f"{col}22",
            (perc, categ, "T1", "files"): f"{col}23",
        }

    ws = wb.create_sheet("expe_1")

    # Titles
    ws["A4"] = "DEPLOY"
    ws["A5"] = "T0"
    ws["A6"] = "T1"
    ws["A7"] = "UPDATE"
    ws["A8"] = "T0"
    ws["A9"] = "T1"
    ws["B3"] = "Mjuz"
    ws["C3"] = "Sync"
    ws["D3"] = "Async"
    ws["E3"] = "Mjuz"
    ws["F3"] = "Sync"
    ws["G3"] = "Async"
    ws["H3"] = "Mjuz"
    ws["I3"] = "Sync"
    ws["J3"] = "Async"
    ws["K3"] = "Mjuz"
    ws["L3"] = "Sync"
    ws["M3"] = "Async"
    ws["B2"] = "2-5%"
    ws["E2"] = "20-30%"
    ws["H2"] = "50-60%"
    ws["K2"] = "100%"

    fill_ws_from_results(case_mapping, data_path, expe_name, ws, ["asynchronous", "synchronous", "mjuz"], "tab1")


def fill_ws_from_results(case_mapping, data_path, expe_name, ws, categ_dirs_list, tab_name):
    # For each category put results
    result_file_name = "global_results_expes_computed.yaml"
    for categ_file in categ_dirs_list:
        with open(f"{data_path}/{expe_name}/{categ_file}/{result_file_name}") as f:
            tab_values = yaml.safe_load(f)[tab_name]

        # Refacto avec compute_globals_results_new.py
        for perc, perc_values in tab_values.items():
            for categ, categ_values in perc_values.items():
                for trans, trans_values in categ_values.items():
                    for metric, metric_values in trans_values.items():
                        if len(metric_values) > 0:
                            mapped_case = case_mapping(perc, categ).get((perc, categ, trans, metric))
                            if mapped_case is not None:
                                col = mapped_case[0]
                                lig = int(mapped_case[1:])

                                if metric == "global_finished_reconf":
                                    ws[mapped_case] = all(metric_values)
                                elif metric == "files":
                                    ws[mapped_case] = len(metric_values)
                                else:
                                    ws[mapped_case] = metric_values["mean"]
                                    ws[col + str(lig + 8)] = metric_values["std"]


def generate_sheet_expe_2(wb, expe_name, data_path):
    col_mapping = {
        "2-2": {"0": "B", "0.5": "D", "1": "F"},
        "25-35": {"0": "G", "0.5": "I", "1": "K"},
        "50-50": {"0": "L", "0.5": "N", "1": "P"},
    }
    def case_mapping(perc, categ):
        col = col_mapping.get(perc, {}).get(categ, None)
        if col is None:
            return {}
        return {
            (perc, categ, "T0", "max_reconf_time"): f"{col}4",
            (perc, categ, "T0", "max_sleeping_time"): f"{col}5",
            (perc, categ, "T0", "max_execution_time"): f"{col}6",

            (perc, categ, "T1", "max_reconf_time"): f"{col}8",
            (perc, categ, "T1", "max_sleeping_time"): f"{col}9",
            (perc, categ, "T1", "max_execution_time"): f"{col}10",

            (perc, categ, "T0", "global_finished_reconf"): f"{col}20",
            (perc, categ, "T0", "files"): f"{col}21",
            (perc, categ, "T1", "global_finished_reconf"): f"{col}22",
            (perc, categ, "T1", "files"): f"{col}23",
        }

    ws = wb.create_sheet("expe_2")

    # Titles
    ws["A3"] = "T0"
    ws["A4"] = "max uptime duration"
    ws["A5"] = "max Sleeping duration"
    ws["A6"] = "max reconf duration"
    ws["A7"] = "T1"
    ws["A8"] = "max uptime duration"
    ws["A9"] = "max Sleeping duration"
    ws["A10"] = "max reconf duration"

    ws["B1"] = "2%"
    ws["G1"] = "25%"
    ws["L1"] = "50%"

    ws["B2"] = "timeout=0%"
    ws["D2"] = "timeout=50%"
    ws["F2"] = "timeout=100%"
    ws["G2"] = "timeout=0%"
    ws["I2"] = "timeout=50%"
    ws["K2"] = "timeout=100%"
    ws["L2"] = "timeout=0%"
    ws["N2"] = "timeout=50%"
    ws["P2"] = "timeout=100%"

    ws["B3"] = "seconds"
    ws["C3"] = "gain (%)"
    ws["D3"] = "seconds"
    ws["E3"] = "gain (%)"
    ws["F3"] = "seconds"
    ws["G3"] = "seconds"
    ws["H3"] = "gain (%)"
    ws["I3"] = "seconds"
    ws["J3"] = "gain (%)"
    ws["K3"] = "seconds"
    ws["L3"] = "seconds"
    ws["M3"] = "gain (%)"
    ws["N3"] = "seconds"
    ws["O3"] = "gain (%)"
    ws["P3"] = "seconds"

    ws["B7"] = "seconds"
    ws["C7"] = "gain (%)"
    ws["D7"] = "seconds"
    ws["E7"] = "gain (%)"
    ws["F7"] = "seconds"
    ws["G7"] = "seconds"
    ws["H7"] = "gain (%)"
    ws["I7"] = "seconds"
    ws["J7"] = "gain (%)"
    ws["K7"] = "seconds"
    ws["L7"] = "seconds"
    ws["M7"] = "gain (%)"
    ws["N7"] = "seconds"
    ws["O7"] = "gain (%)"
    ws["P7"] = "seconds"

    fill_ws_from_results(case_mapping, data_path, expe_name, ws, ["asynchronous"], "tab2")

    # Fill gains
    gain_col_mapping = {
        "2-2": {"0": "C", "0.5": "E"},
        "25-35": {"0": "H", "0.5": "J"},
        "50-50": {"0": "M", "0.5": "O"},
    }
    def gain_case_mapping(perc, categ):
        col = gain_col_mapping.get(perc, {}).get(categ, None)
        if col is None:
            return {}
        return {
            (perc, categ, "T0", "max_reconf_time"): f"{col}4",
            (perc, categ, "T0", "max_sleeping_time"): f"{col}5",
            (perc, categ, "T0", "max_execution_time"): f"{col}6",

            (perc, categ, "T1", "max_reconf_time"): f"{col}8",
            (perc, categ, "T1", "max_sleeping_time"): f"{col}9",
            (perc, categ, "T1", "max_execution_time"): f"{col}10",
        }

    def compute_gain(begin, end):
        return round(begin-end)*100/end

    for perc in ["2-2", "25-35", "50-50"]:
        for categ in ["0", "0.5"]:
            for parameters, output_case in gain_case_mapping(perc, categ).items():
                end_perc, end_categ, end_trans, end_metric = parameters
                map_b = ws[case_mapping(perc, categ)[parameters]].value
                if map_b:
                    begin_case = float(map_b.replace(",", "."))
                    end_case = float(ws[case_mapping(perc, "1")[(end_perc, "1", end_trans, end_metric)]].value.replace(",", "."))
                    result_gain = compute_gain(begin_case, end_case)
                    ws[output_case] = round(result_gain, 2)


wb = Workbook()
expe_name = "mascots"

# Generate sheets
generate_sheet_expe_1(wb, expe_name, "/home/aomond/experiments_results/concerto-d/prod")
# generate_sheet_expe_2(wb, expe_name, "/home/aomond/experiments_results/concerto-d/prod")

# Save the file
wb.save(f"{expe_name}-prod-new-uptimes-2-2.xlsx")
